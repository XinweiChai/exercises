import openpyxl
from datetime import datetime
from datetime import timedelta
import pandas as pd

holiday = [1, 2, 3, 4, 5, 10, 16, 17, 23, 24, 30, 31]

y_offset = 7


def work_statistics(fn):
    workbook = openpyxl.load_workbook(fn)
    sheet = workbook.worksheets[0]
    rows = (sheet.max_row - y_offset) // 2
    days = sheet.max_column
    for i in range(rows):
        absence_count = 0
        not_enough_count = 0
        absence = []
        not_enough = []
        name = x = sheet.cell(i * 2 + y_offset - 1, 11).value
        for j in range(1, days + 1):
            if j in holiday:
                continue
            x = sheet.cell(i * 2 + y_offset, j).value
            if not x or len(x) < 10:
                absence_count += 1
                absence.append(j)
                continue
            start = datetime.strptime(x[0:5], "%H:%M")
            end = datetime.strptime(x[-5:], "%H:%M")
            if (end - start).seconds < 3600:
                absence_count += 1
                absence.append(j)
            elif (end - start).seconds < 9 * 3600 - 6 * 60:
                not_enough_count += 1
                not_enough.append(j)
        absence_count, absence, not_enough_count, not_enough = modify_res(absence_count, absence, not_enough_count,
                                                                          not_enough, name)
        sheet.cell(i * 2 + y_offset, days + 1, absence_count)
        sheet.cell(i * 2 + y_offset, days + 2, absence)
        sheet.cell(i * 2 + y_offset, days + 3, not_enough_count)
        sheet.cell(i * 2 + y_offset, days + 4, not_enough)
    sheet.cell(1, days + 1, "缺打卡天数")
    sheet.cell(1, days + 2, "缺打卡日期")
    sheet.cell(1, days + 3, "上班时间不足天数")
    sheet.cell(1, days + 4, "上班时间不足日期")
    workbook.save("res.xlsx")


def modify_res(ab_c, ab, ne, ne_c, name):
    df1 = pd.DataFrame(pd.read_excel("钉钉出差.xlsx"))
    df2 = pd.DataFrame(pd.read_excel("钉钉外出.xlsx"))
    df3 = pd.DataFrame(pd.read_excel("钉钉请假.xlsx"))
    dfs = [df1, df2, df3]
    for df in dfs:
        for idx, i in df.iterrows():
            if name == i["标题"][:-5]:
                begin_time = datetime.strptime(i["开始时间"], "%Y-%m-%d")
                end_time = datetime.strptime(i["结束时间"], "%Y-%m-%d")
                days = (end_time - begin_time).days + 1
                period = [(begin_time + timedelta(days=day)).day for day in range(days)]
                ab_c -= len(period)
                for k in period:
                    ab.remove(k)
    return ab_c, ab, ne, ne_c


def count_days(fn, workbook2):
    workbook = openpyxl.load_workbook(fn)
    sheet1 = workbook.worksheets[0]
    sheet2 = workbook2.worksheets[0]
    pos_header = 0
    pos_days = 0
    pos_type = 0
    pos_reason = 0
    for i in range(1, sheet1.max_column + 1):
        if sheet1.cell(1, i).value == "标题":
            pos_header = i
        elif "天数" in sheet1.cell(1, i).value:
            pos_days = i
        elif "事由" in sheet1.cell(1, i).value:
            pos_reason = i
        elif "请假类型" in sheet1.cell(1, i).value:
            pos_type = i

    rows1 = sheet1.max_row - 1
    rows2 = sheet2.max_row - 1
    for i in range(2, rows1 + 2):
        if "出差" in fn:
            x_offset = 13
        else:
            if sheet1.cell(i, pos_type).value == "年假":
                x_offset = 8
            elif sheet1.cell(i, pos_type).value == "事假":
                x_offset = 3
            elif sheet1.cell(i, pos_type).value == "其他":
                x_offset = 14
        name = sheet1.cell(i, pos_header).value
        name = name[:-5]
        days = float(sheet1.cell(i, pos_days).value)
        reason = sheet1.cell(i, pos_reason).value
        for j in range(5, rows2 + 1):
            if sheet2.cell(j, 2).value == name:
                if not sheet2.cell(j, 17).value:
                    sheet2.cell(j, 17).value = reason + ";"
                else:
                    sheet2.cell(j, 17).value += reason + ";"
                if not sheet2.cell(j, x_offset).value:
                    sheet2.cell(j, x_offset).value = days
                else:
                    sheet2.cell(j, x_offset).value += days
                break
    return workbook2


def count_exceptions():
    workbook2 = openpyxl.load_workbook("考勤汇总表-最终.xlsx")
    workbook2 = count_days("钉钉出差.xlsx", workbook2)
    workbook2 = count_days("钉钉请假.xlsx", workbook2)
    workbook2.save("final.xlsx")


if __name__ == "__main__":
    work_statistics("考勤机原始考勤.xlsx")
    count_exceptions()
