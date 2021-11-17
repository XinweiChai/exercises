import openpyxl
from datetime import datetime
import easygui

default_offset = 6


def work_statistics(fn, y_offset=default_offset, holidays=None, sheet=0):
    workbook = openpyxl.load_workbook(fn)
    sheet = workbook.worksheets[sheet]
    rows = (sheet.max_row - y_offset) // 2
    days = sum(sheet.cell(y_offset - 2, i).value is not None for i in range(1, sheet.max_column + 1))
    for i in range(rows):
        absence_count = 0
        not_enough_count = 0
        absence = []
        not_enough = []
        for j in range(1, days + 1):
            if j in holidays:
                continue
            x = sheet.cell(i * 2 + y_offset, j).value
            # print(x)
            if not x or len(x) < 10:
                absence_count += 1
                absence.append(str(j))
                continue
            start = datetime.strptime(x[0:5], "%H:%M")
            end = datetime.strptime(x[-5:], "%H:%M")
            if (end - start).seconds < 3600:
                absence_count += 1
                absence.append(str(j))
            # elif (end - start).seconds < 9 * 3600 - 6 * 60:
            elif start > datetime.strptime('08:35', '%H:%M') or end < datetime.strptime('17:30', '%H:%M'):
                not_enough_count += 1
                not_enough.append(str(j))
        sheet.cell(i * 2 + y_offset, days + 1, absence_count)
        sheet.cell(i * 2 + y_offset, days + 2, ' '.join(absence))
        sheet.cell(i * 2 + y_offset, days + 3, not_enough_count)
        sheet.cell(i * 2 + y_offset, days + 4, ' '.join(not_enough))
    sheet.cell(1, sheet.max_column + 1, "缺打卡天数")
    sheet.cell(1, sheet.max_column + 1, "缺打卡日期")
    sheet.cell(1, sheet.max_column + 1, "上班时间不足天数")
    sheet.cell(1, sheet.max_column + 1, "上班时间不足日期")
    workbook.save("res.xlsx")


def openfile():
    sheet = int(easygui.enterbox('Sheet', default='0'))
    offset = easygui.enterbox('Y-offset', default='6')
    holidays = easygui.enterbox('Holidays split by space', default='1 2 3 4 5 10 16 17 23 24 30 31')
    if holidays:
        holidays = set(map(int, holidays.split()))
    if offset and offset.isdigit():
        offset = int(offset) + 1
    else:
        offset = default_offset
    fn = easygui.fileopenbox()
    if fn and holidays:
        work_statistics(fn, offset, holidays, sheet)


if __name__ == "__main__":
    # work_statistics("考勤机原始考勤.xlsx")
    openfile()
