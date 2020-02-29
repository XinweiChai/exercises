import openpyxl
import random

afternoon_bias = 0.3
xiong_min = 35.9
chai_min = 35.7
num_spaces = 96
morning_column = 3
afternoon_column = 4
evening_column = 5
xiong_line = 5
chai_line = 6

def rand_excel(fn):
    workbook = openpyxl.load_workbook(fn)
    sheet = workbook.worksheets[0]
    for k in range(1, 32):
        sheet.cell(2, 1, num_spaces * ' ' + "日期：2020年3月" + str(k) + "日")
        for j in [morning_column, afternoon_column, evening_column]:
            sheet.cell(xiong_line, j, xiong_min + round(random.random() / 3, 1))
            sheet.cell(chai_line, j, chai_min + round(random.random() / 3, 1))
        sheet.cell(xiong_line, afternoon_column, xiong_min + afternoon_bias + random.randint(0, 3) / 10)
        sheet.cell(chai_line, afternoon_column, chai_min + afternoon_bias + random.randint(0, 3) / 10)
        workbook.save("C:/Users/Administrator/Desktop/body_temperature/202003" + str(k).zfill(2) + ".xlsx")


if __name__ == "__main__":
	rand_excel("C:/Users/Administrator/Desktop/body_temperature/20200331.xlsx")